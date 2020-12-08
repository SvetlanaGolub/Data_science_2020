import csv
import os
from openpyxl import load_workbook

full_table = []
files = os.listdir(path="./YourBoardingPassDotAero")
for file in files:
    wb = load_workbook("./YourBoardingPassDotAero/" + file)
    sheet = wb.active
    table = {'Name': sheet.cell(row=3, column=2).value, 'Sex': sheet.cell(row=3, column=1).value,
             'From': sheet.cell(row=5, column=4).value, 'To': sheet.cell(row=5, column=8).value,
             'Date': sheet.cell(row=9, column=1).value, 'Time': sheet.cell(row=9, column=3).value,
             'Flight_number': sheet.cell(row=5, column=1).value,
             'E-ticket': sheet.cell(row=13, column=5).value, 'Class': sheet.cell(row=3, column=8).value,
             'Sequence': sheet.cell(row=1, column=8).value,
             'Seat': sheet.cell(row=11, column=8).value, 'PNR': sheet.cell(row=13, column=2).value}
    if str(table['Name']).split()[0].endswith(("OVA", "EVA", "IVA", "AVA", "UVA", "INA", "OV", "EV", "IV", "AV", "UV", "IN")):
        full_name = table['Name'].split()
        if len(full_name) == 2:
            table['Name'] = full_name[1] + ' ' + full_name[0]
        elif len(full_name[1]) == 1:
            table['Name'] = full_name[2] + ' ' + full_name[0]
        elif len(full_name[2]) == 1:
            table['Name'] = full_name[1] + ' ' + full_name[0]
    full_table.append(table)

column_names = full_table[0].keys()
with open('BoardingPass.csv', 'a', newline='') as output_file:
    dict_writer = csv.DictWriter(output_file, column_names)
    dict_writer.writeheader()
    dict_writer.writerows(full_table)
